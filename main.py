import pandas as pd 
import os
from openpyxl import load_workbook

def readFile(path):
    
    df = pd.ExcelFile(path)
    hoja= df.sheet_names[0]

    wb = load_workbook(path)
    ws = wb[hoja]

    mapping = {}

    for entry, data_boundary in ws.tables.items():
        data = ws[data_boundary]
        content = [ [cell.value for cell in ent] for ent in data ]
        header = content[0]
        rest = content[1:]
        df = pd.DataFrame(rest, columns = header)
        mapping[entry] = df
        dates=[]
        for date in mapping['tabla']['Semana tentativa para finalizar']:
            dates.append(date)
        return dates

def formatearFechas(fechas):
    fechasLimpias = []
    fechasRevisar = []

    for f in fechas:

        if f == "2 de mayo" or f == "\xa02 de mayo":
            fechasLimpias.append("2 de mayo")
            continue

        if f == "\xa0Primer semana de mayo\n" or f == "\xa0\xa0Primer semana de mayo":
            fechasLimpias.append("primer semana de mayo")

        if f == "PRIMERA O SEGUNDA SEMANA DE MAYO\n" or f == "\xa0PRIMERA O SEGUNDA SEMANA DE MAYO":
            fechasLimpias.append("primera o segunda semana de mayo")
            continue

        if f == "SEGUNDA SEMANA DE MAYO" or f == "Segunda semana de Mayo" or f == "\xa0Segunda semana de Mayo" or f == "segunda semana de mayo":
            fechasLimpias.append("segunda semana de mayo")
            continue

        if f == "Tercera semna mayo" or f == "tercera semana de mayo":
            fechasLimpias.append("Tercera semana de mayo")
            continue

        if f == "\xa0DE 04 AL 10 DE MAYO":
            fechasLimpias.append("04 AL 10 DE MAYO")
            continue

        if f == "\xa015 a 20 mayo" or f == "15 a 20 mayo":
            fechasLimpias.append("15 a 20 mayo")
            continue

        if f == "\xa0Durante la primera quincena del mes de mayo":
            fechasLimpias.append("Durante la primera quincena del mes de mayo")
            continue

        if f == "\xa08 mayo":
            fechasLimpias.append("8 mayo")
            continue

        if f == "\xa011 mayo":
            fechasLimpias.append("11 mayo")

        if f == "\xa011-14 mayo" or f == "11 mayo - 14 mayo" or f == "11-14 Mayo":
            fechasLimpias.append("11-14 mayo")
            continue

        if f == "11 - 15 mayo" or f == "\xa011-15 de Mayo" or f == "\xa0Del 11 al 15 de mayo." or f == "\xa011-15 de mayo" or f == "11-15 de mayo" or f == "11 al 15  Mayo" or f == "11 al 15 de Mayo":
            fechasLimpias.append("11 - 15 mayo")
            continue

        # Separado de las anteriores, por si el año es diferente
        if f == "\xa011 a 15 de mayo 2020":
            fechasLimpias.append("11 a 15 de mayo 2020")
            continue

        if f == "\xa04-7 de Mayo":
            fechasLimpias.append("4-7 de Mayo")
            continue

        if f == "4 - 8 mayo" or f == "4-8 MAYO":
            fechasLimpias.append("4 - 8 mayo")
            continue

        if f == "\xa04-9 DE MAYO":
            fechasLimpias.append("4-9 DE MAYO")
            continue

        if f == "11-14 de Mayo" or f == "11 a 14 de Mayo":
            fechasLimpias.append("11-14 de Mayo")
            continue

        if f == "11-14 de Mayo de 2020":
            fechasLimpias.append("11-14 de Mayo de 2020")
            continue

        if f == "\xa012-16 MAYO":
            fechasLimpias.append("12-16 MAYO")
            continue

        if f == "18 mayo - 21 mayo ":
            fechasLimpias.append("18 - 21 mayo")

        if f == "18 - 22 mayo" or f == "\xa018 al 22 de mayo" or f == "18-22 de mayo" or f == "18 al 22  Mayo" or f == "18 al 22 de Mayo":
            fechasLimpias.append("18 - 22 mayo")
            continue

        if f == "18-22 de Mayo 2020":
            fechasLimpias.append("18-22 de Mayo 2020")
            continue

        if f == "18 al 23 de  Mayo" or f == "18 al  23 de Mayo ":
            fechasLimpias.append("18 al 23 de  Mayo")
            continue

        if f == "\xa022-29 de mayo":
            fechasLimpias.append("22-29 de mayo")
            continue

        if f == "25 de mayo 2020":
            fechasLimpias.append("25 de mayo 2020")
            continue

        if f == "25 al 29 mayo" or f == "25-29 de Mayo" or f == "25 al 29 de Mayo ":
            fechasLimpias.append("25 al 29 mayo")
            continue

        if f == "4-6 DE MAYO" or f == "\xa04-6 DE MAYO":
            fechasLimpias.append("4-6 DE MAYO")
            continue

        if f == "\xa04-7 mayo" or f == "\xa04-7 de Mayo":
            fechasLimpias.append("4-7 mayo")
            continue

        if f == "Semana del 04 al 08 de Mayo de 2020":
            fechasLimpias.append("Semana del 04 al 08 de Mayo de 2020")
            continue

        if f == "Semana del 11 al 15 de Mayo de 2020":
            fechasLimpias.append("Semana del 11 al 15 de Mayo de 2020")
            continue

        if f == "04-07 Mayo 2020":
            fechasLimpias.append("04-07 Mayo 2020")
            continue

        if f == "\xa07 de mayo " or f == "7 de mayo":
            fechasLimpias.append("7 de mayo")
            continue

        if f == "11-14 Mayo 2020" or f == "11-14 Mayo/2020":
            fechasLimpias.append("11-14 Mayo 2020")
            continue

        if f == "Semana del 11 al 14 de Mayo de 2020":
            fechasLimpias.append("Semana del 11 al 14 de Mayo de 2020")
            continue

        if f == "11-15 Mayo":
            fechasLimpias.append("11-15 Mayo")
            continue

        if f == "11-15 Mayo 2020":
            fechasLimpias.append("11-15 Mayo 2020")
            continue

        if f == "\xa018-21 mayo":
            fechasLimpias.append("18-21 mayo")
            continue

        if f == "18-21 mayo 2020":
            fechasLimpias.append("18-21 mayo 2020")
            continue

        if f == "18-22 Mayo 2020":
            fechasLimpias.append("18-22 Mayo 2020")
            continue

        if f == "23 y 30 mayo 2020":
            fechasLimpias.append("23 y 30 mayo 2020")
            continue

        if f == "semana del 29 de mayo":
            fechasLimpias.append("semana del 29 de mayo")
            continue

        if f == "24 - 30 mayo" or f == "24 AL 30 DE MAYO":
            fechasLimpias.append("24 - 30 mayo")
            continue

        if f == "25-31 de mayo de 2020":
            fechasLimpias.append("25-31 de mayo de 2020")
            continue

        if f == "04 AL 08 ABRIL":
            fechasLimpias.append("04 AL 08 ABRIL")
            continue

        if f == "\xa027-30 de abril" or f == "\xa027 – 30 Abril" or f == "27-30 de Abril":
            fechasLimpias.append("27-30 de abril")
            continue

        if f == "\xa0Última semana de mayo \n" or f == "Última semana de mayo" or f == "Última semana de mayo " or f == "Última semana de  mayo ":
            fechasLimpias.append("Última semana de mayo")

        # Abril
        if f == "\xa023 a 30 abril" or f == "\xa0\xa023 a 30 abril":
            fechasLimpias.append("23 a 30 abril")
            continue

        if f == "\xa027 de abril":
            fechasLimpias.append("27 de abril")
            continue

        if f == "27-30 abril\xa0" or f == "27 AL 30 ABRIL":
            fechasLimpias.append("27-30 abril")
            continue

        if f == "27-30 abril o primera semana de mayo\n":
            fechasLimpias.append("27-30 abril o primera semana de mayo")

        if f == "\xa0Esta ya termina a fin de abril ":
            fechasLimpias.append("Esta ya termina a fin de abril")
            continue

        if f == "\xa0DEL 17 AL 22 DE MAYO\n":
            fechasLimpias.append("DEL 17 AL 22 DE MAYO")
            continue

        if f == "23 – 30 de abril":
            fechasLimpias.append("23 – 30 de abril")
            continue

        if f == "\xa0Tercera semana de mayo":
            fechasLimpias.append("Tercera semana de mayo")
            continue

        # Junio
        if f == "Primera  semana de Junio ":
            fechasLimpias.append("Primera  semana de Junio")
            continue
        
        if f == "\xa08 – 13 de junio de 2020":
            fechasLimpias.append("8 – 13 de junio de 2020")
            continue

        if f == "1 - 5 junio" or f == "01 - 05 junio":
            fechasLimpias.append("1 - 5 junio")
            continue

        if f == "15 de junio":
            fechasLimpias.append("15 de junio")
            continue

        if f == "09-2015 Junio":
            fechasLimpias.append("09-2015 Junio")
            continue

        if f == "MEDIADOS DE JUNIO":
            fechasLimpias.append("MEDIADOS DE JUNIO")
            continue
        
        # Fechas que no entiendo
        if f == "\xa030-4-20":
            fechasLimpias.append("30-4-20")
            continue

        if f == "\xa027 al 01 de mayo" or f == "\xa0\xa027 al 01 de mayo":
            fechasLimpias.append("27 al 01 de mayo")
            continue

        if f == "se finaliza la ultima semana de abril 2020.La recuperacion sera del 27 al 29 de abril.\n":
            fechasLimpias.append("ultima semana de abril 2020.La recuperacion sera del 27 al 29 de abril.")
            continue

        if f == "Sin tentativa \nSegún acuerdos\nde la UNAH.\n\n":
            fechasLimpias.append("Sin tentativa")
            continue

        if f == "12-2016 mayo":
            fechasLimpias.append("12-2016 mayo")
            continue

        if f == "4-2008 mayo":
            fechasLimpias.append("4-2008 mayo")
            continue

        if f == "15-2020 mayo":
            fechasLimpias.append("15-2020 mayo")
            continue

        if f == "18-2023 Mayo" or f == "18-2023 mayo":
            fechasLimpias.append("18-2023 Mayo")
            continue

        if f == "25-2029 mayo":
            fechasLimpias.append("25-2029 mayo")
            continue

        if f == "18-2022 mayo":
            fechasLimpias.append("18-2022 mayo")
            continue

        if f == "8-2012 junio":
            fechasLimpias.append("8-2012 junio")
            continue

        if f == "11-2015 mayo":
            fechasLimpias.append("11-2015 mayo")
            continue

        if f == "12-2016 mayo":
            fechasLimpias.append("12-2016 mayo")
            continue

        if f == "DEL 8 AL 12 DE JUNIO":
            fechasLimpias.append("8 AL 12 DE JUNIO")
            continue

        if f == "12-2016 junio":
            fechasLimpias.append("12-2016 junio")
            continue

        if f == "22-2026 Junio":
            fechasLimpias.append("22-2026 Junio")
            continue

        if f == "Sem. 14":
            fechasLimpias.append("Sem. 14")
            continue

        if f == "Sem. 16":
            fechasLimpias.append("Sem. 16")
            continue

        # Fechas Invalidas
        if f == "datetime.datetime(2020, 5, 11, 0, 0)":
            fechasRevisar.append("datetime.datetime(2020, 5, 11, 0, 0)")
            continue

        if f == "datetime.datetime(2020, 5, 16, 0, 0)":
            fechasRevisar.append("datetime.datetime(2020, 5, 16, 0, 0)")
            continue

        if f == " datetime.datetime(2020, 4, 30, 0, 0)":
            fechasRevisar.append(" datetime.datetime(2020, 4, 30, 0, 0)")
            continue

    return fechasLimpias

def voto(fechas):
    return fechas[1]

def contador(fechasLimpias, fechasOrdenadas):
    maximo = []

    for elemento in fechasOrdenadas:
        temp = [elemento, fechasLimpias.count(elemento)]
        maximo.append(temp)

    maximo.sort(key=voto, reverse=True)
    return maximo

def main():
    paths='C:/Users/jesus/OneDrive/Documentos/Tarea_BigData/Data/'
    dates=[]

    for path , dir, files in os.walk(paths):
        for f in files:
            p= os.path.join(path,f)
            dates+=readFile(p)
    
    print("Total de Fechas: ", len(dates))
    print("Fechas Unicas: ", len(set(dates)))
    print("\n")
    print("Despues de limpiar...\n")
    fechasLimpias = formatearFechas(dates)
    print("Total de Fechas: ", len(dates))
    print("Fechas Unicas: ", len(set(fechasLimpias)))
    finalizar = contador(fechasLimpias, list(set(fechasLimpias)))

    print("La fecha tentativa de fin de periodo es: ", finalizar[0][0], "con la cantidad de ", str(finalizar[0][1]) + " votos")

main()